from openpyxl import Workbook

from schedule_parser.rule_resolver import parse_direct_time_cell, parse_inline_override, parse_legend_from_sheet, resolve_shift_code
from schedule_parser.schema import ParsedShiftTime
from schedule_parser.time_utils import parse_time_range


def _shift(code, start, end, hours):
    return ParsedShiftTime(
        code=code,
        start=start,
        end=end,
        hours=hours,
        source="legend",
        source_text=f"{code} {start}-{end}",
        source_cell="A1",
        applies_to="default",
        specific_dates=[],
    )


def test_direct_time_cell_uses_adjacent_hours_when_valid():
    result = parse_direct_time_cell("9:45 - 20:15", adjacent_hours=10.25)

    assert result["scheduled_in"] == "09:45"
    assert result["scheduled_out"] == "20:15"
    assert result["scheduled_hours"] == 10.25
    assert result["resolution_source"] == "direct_time"


def test_time_range_skips_leading_legend_hours_prefix():
    result = parse_time_range("A 10.5 9:45am-6:45pm (9 hr) (tentative lunch: 1-2pm)")

    assert result["start"] == "09:45"
    assert result["end"] == "18:45"
    assert result["hours"] == 9


def test_time_range_uses_duration_to_disambiguate_am_pm():
    result = parse_time_range("A1 6.5 1:45am-6:45pm (5 hr)")

    assert result["start"] == "13:45"
    assert result["end"] == "18:45"
    assert result["hours"] == 5


def test_inline_override_forms():
    shifts = {
        "D": _shift("D", "09:45", "20:15", 10.5),
        "B": _shift("B", "10:45", "20:15", 9.5),
        "A1": _shift("A1", "13:45", "18:45", 5.0),
    }

    assert parse_inline_override("D (2pm - 8:15pm)", shifts)["scheduled_in"] == "14:00"
    till = parse_inline_override("B (till 15:45)", shifts)
    assert till["scheduled_in"] == "10:45"
    assert till["scheduled_out"] == "15:45"
    a1 = parse_inline_override("A1(13:45-5:45PM)", shifts)
    assert a1["scheduled_in"] == "13:45"
    assert a1["scheduled_out"] == "17:45"


def test_shift_code_with_digit_prefers_legend_over_inline_override():
    shifts = {
        "B": _shift("B", "10:45", "20:15", 9.5),
        "B1": _shift("B1", "13:45", "20:15", 6.5),
    }

    result = resolve_shift_code("B1", "2025-04-14", shifts)

    assert parse_inline_override("B1", shifts) is None
    assert result["shift_code"] == "B1"
    assert result["scheduled_in"] == "13:45"
    assert result["scheduled_out"] == "20:15"
    assert result["scheduled_hours"] == 6.5
    assert result["resolution_source"] == "legend"


def test_unknown_shift_code_is_preserved_with_warning():
    shifts = {
        "A": _shift("A", "09:45", "18:45", 9.0),
        "C": _shift("C", "09:45", "15:45", 6.0),
    }

    unknown = resolve_shift_code("Z9", "2025-04-01", shifts)
    assert unknown["resolution_source"] == "unresolved"
    assert "UNKNOWN_SHIFT_CODE" in unknown["warnings"]


def test_or_shift_equivalent_no_warning():
    shifts = {
        "A": _shift("A", "09:45", "18:45", 9.0),
        "C": _shift("C", "09:45", "18:45", 9.0),
    }

    result = resolve_shift_code("A/C", "2025-04-01", shifts)

    assert "AMBIGUOUS_SHIFT_CODE" not in result["warnings"]
    assert result["resolution_source"] == "or_equivalent"
    assert result["scheduled_in"] == "09:45"
    assert result["scheduled_out"] == "18:45"
    assert result["scheduled_hours"] == 9.0
    assert [option.code for option in result["shift_options"]] == ["A", "C"]


def test_or_shift_uses_direct_combined_code_if_present():
    shifts = {
        "A/C": _shift("A/C", "09:45", "18:45", 9.0),
        "A": _shift("A", "09:45", "18:45", 9.0),
        "C": _shift("C", "13:45", "18:45", 5.0),
    }

    result = resolve_shift_code("A/C", "2025-04-01", shifts)

    assert result["resolution_source"] == "legend_combined_code"
    assert result["scheduled_in"] == "09:45"
    assert result["scheduled_out"] == "18:45"
    assert result["scheduled_hours"] == 9.0
    assert "AMBIGUOUS_SHIFT_CODE" not in result["warnings"]


def test_or_shift_different_options_no_warning():
    shifts = {
        "A": _shift("A", "09:45", "18:45", 9.0),
        "C": _shift("C", "13:45", "18:45", 5.0),
    }

    result = resolve_shift_code("A/C", "2025-04-01", shifts)

    assert "AMBIGUOUS_SHIFT_CODE" not in result["warnings"]
    assert result["warnings"] == []
    assert result["resolution_source"] == "or_options"
    assert result["scheduled_in"] == ""
    assert result["scheduled_out"] == ""
    assert result["scheduled_hours"] is None
    assert [(option.code, option.start, option.end, option.hours) for option in result["shift_options"]] == [
        ("A", "09:45", "18:45", 9.0),
        ("C", "13:45", "18:45", 5.0),
    ]


def test_or_shift_unknown_part_warning():
    shifts = {"A": _shift("A", "09:45", "18:45", 9.0)}

    result = resolve_shift_code("A/Z9", "2025-04-01", shifts)

    assert result["shift_code"] == "A/Z9"
    assert result["resolution_source"] == "unresolved_or_options"
    assert "UNKNOWN_OR_SHIFT_PART" in result["warnings"]
    assert [(option.code, option.start, option.end, option.hours) for option in result["shift_options"]] == [
        ("A", "09:45", "18:45", 9.0),
    ]


def test_legend_parser_ignores_need_and_total_summary_rows():
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1, value="Name")
    sheet.cell(2, 1, value="Staff One")
    sheet.cell(4, 1, value="Need 5 7 7 4 5 5 5 5 7 7 2564")
    sheet.cell(5, 1, value="Total 5 6 7 4 0 5 5 5 7 7")
    sheet.cell(6, 1, value="A 9:45am-6:45pm (9 hr)")

    shifts = parse_legend_from_sheet(sheet, sheet, header_row=1, staff_end_row=2)

    assert "A" in shifts
    assert "N" not in shifts
    assert "T" not in shifts
