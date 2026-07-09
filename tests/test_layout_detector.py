from datetime import date, timedelta

from openpyxl import Workbook

from schedule_parser.excel_inspector import inspect_workbook
from schedule_parser.layout_detector import (
    DIRECT_TIME_WITH_HOURS_COLUMNS,
    FIRST_SHEET_SELECTED_OVER_HIGHER_SCORING_SHEET,
    SHIFT_CODE_MATRIX_WITH_LEGEND,
    detect_schedule_layout,
)

from conftest import save_apr_style_workbook, save_jan_style_workbook


def test_detects_direct_time_with_hours_columns(tmp_path):
    inspected = inspect_workbook(save_jan_style_workbook(tmp_path / "jan.xlsx"))

    result = detect_schedule_layout(inspected)

    assert result.sheet_name == "Jan"
    assert result.header_row == 6
    assert result.layout_type == DIRECT_TIME_WITH_HOURS_COLUMNS
    assert [col["letter"] for col in result.likely_date_columns] == ["E", "G", "I"]


def test_detects_shift_code_matrix_with_legend(tmp_path):
    inspected = inspect_workbook(save_apr_style_workbook(tmp_path / "apr.xlsx"))

    result = detect_schedule_layout(inspected)

    assert result.sheet_name == "Apr"
    assert result.header_row == 6
    assert result.layout_type == SHIFT_CODE_MATRIX_WITH_LEGEND
    assert result.diagnostics["date_count"] == 30


def test_prefers_first_parseable_sheet_over_higher_scoring_old_sheet(tmp_path):
    path = tmp_path / "multi_sheet.xlsx"
    wb = Workbook()
    current = wb.active
    current.title = "1 - 30 Jun "
    old = wb.create_sheet("Old")
    _fill_shift_matrix_sheet(current, date_count=5, staff_count=2)
    _fill_shift_matrix_sheet(old, date_count=10, staff_count=10)
    wb.save(path)

    inspected = inspect_workbook(path)
    result = detect_schedule_layout(inspected)

    assert result.sheet_name == "1 - 30 Jun "
    assert result.layout_type == SHIFT_CODE_MATRIX_WITH_LEGEND
    assert FIRST_SHEET_SELECTED_OVER_HIGHER_SCORING_SHEET in result.warnings
    ignored = result.diagnostics["ignored_higher_scoring_sheet"]
    assert ignored["sheet_name"] == "Old"
    assert ignored["overall_score"] > result.diagnostics["overall_score"]


def _fill_shift_matrix_sheet(ws, date_count: int, staff_count: int) -> None:
    header_row = 6
    for index, value in enumerate(["Name", "Start Date", "Staff ID", "Mobile"], start=1):
        ws.cell(header_row, index, value=value)
    start_date = date(2026, 6, 1)
    for offset in range(date_count):
        ws.cell(header_row, 5 + offset, value=start_date + timedelta(days=offset))
    for row_offset in range(staff_count):
        row = header_row + 1 + row_offset
        ws.cell(row, 1, value=f"Staff {row_offset + 1}")
        ws.cell(row, 3, value=f"S{row_offset + 1:03d}")
        ws.cell(row, 4, value="91234567")
        for offset in range(date_count):
            ws.cell(row, 5 + offset, value="A")
    ws.cell(header_row + staff_count + 1, 1, value="Total")
    ws.cell(header_row + staff_count + 3, 1, value="A")
    ws.cell(header_row + staff_count + 3, 2, value="9:45am-6:45pm (9 hr)")
