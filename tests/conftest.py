from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook


def save_jan_style_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jan"
    header_row = 6
    headers = ["Name", "Start Date", "Staff ID", "Mobile"]
    for index, value in enumerate(headers, start=1):
        ws.cell(header_row, index, value=value)
    for offset, col in enumerate([5, 7, 9]):
        ws.cell(header_row, col, value=date(2025, 1, offset + 1))
        ws.cell(header_row, col + 1, value="Hours")
    staff_rows = [
        ("John Chan", "2024-01-01", "S001", "91234567", ["9:45 - 20:15", "13:45 - 20:15", "10:00 -20:15"], [10.5, 6.5, 10.25]),
        ("Mary Lee", "2024-02-01", "S002", "92345678", ["13:45 - 20:15", "", "9:45 - 20:15"], [6.5, None, 10.5]),
    ]
    for row_index, (name, start, staff_id, mobile, shifts, hours) in enumerate(staff_rows, start=7):
        ws.cell(row_index, 1, value=name)
        ws.cell(row_index, 2, value=start)
        ws.cell(row_index, 3, value=staff_id)
        ws.cell(row_index, 4, value=mobile)
        for shift, hour, col in zip(shifts, hours, [5, 7, 9]):
            ws.cell(row_index, col, value=shift)
            ws.cell(row_index, col + 1, value=hour)
    ws.cell(9, 1, value="Total")
    wb.save(path)
    return path


def save_apr_style_workbook(path: Path, unknown_code: bool = False) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Apr"
    header_row = 6
    headers = ["Name", "Start Date", "Staff ID", "Mobile"]
    for index, value in enumerate(headers, start=1):
        ws.cell(header_row, index, value=value)
    start = date(2025, 4, 1)
    for offset in range(30):
        ws.cell(header_row, 5 + offset, value=start + timedelta(days=offset))
    for offset, label in enumerate(["Days", "Normal Days Total Hours", "Salary", "Client"], start=35):
        ws.cell(header_row, offset, value=label)

    rows = [
        ("John Chan", "S001", "91234567", ["A", "B", "D (2pm - 8:15pm)", "A1(13:45-5:45PM)"]),
        ("Mary Lee", "S002", "92345678", ["C", "D", "B (till 15:45)", "B1"]),
    ]
    for row_index, (name, staff_id, mobile, codes) in enumerate(rows, start=7):
        ws.cell(row_index, 1, value=name)
        ws.cell(row_index, 3, value=staff_id)
        ws.cell(row_index, 4, value=mobile)
        for offset, code in enumerate(codes):
            ws.cell(row_index, 5 + offset, value=code)
    if unknown_code:
        ws.cell(7, 9, value="Z9")
    ws.cell(9, 1, value="Total")

    ws.cell(11, 1, value="Tue - Sun")
    legend = [
        ("A", "9:45am-6:45pm (9 hr)"),
        ("B", "10:45am-8:15pm (9.5 hr)"),
        ("C", "9:45am-3:45pm (6 hr)"),
        ("D", "9:45am-8:15pm (10.5 hr)"),
        ("E", "12:45pm-4:45pm (4 hr)"),
        ("A1", "1:45am-6:45pm (5 hr)"),
        ("B1", "1:45pm-8:15pm (6.5 hr)"),
    ]
    for offset, (code, text) in enumerate(legend, start=12):
        ws.cell(offset, 1, value=code)
        ws.cell(offset, 2, value=text)
    ws.cell(20, 1, value="Mon")
    ws.cell(21, 1, value="A")
    ws.cell(21, 2, value="1:45pm-6:45pm (5 hr)")
    wb.save(path)
    return path


def save_multi_month_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Apr - May"
    header_row = 6
    headers = ["Name", "Start Date", "Staff ID", "Mobile"]
    for index, value in enumerate(headers, start=1):
        ws.cell(header_row, index, value=value)
    for offset, value in enumerate([
        date(2026, 4, 29),
        date(2026, 4, 30),
        date(2026, 5, 1),
        date(2026, 5, 2),
    ]):
        ws.cell(header_row, 5 + offset, value=value)

    rows = [
        ("Ivy Hung", "S001", "91234567", ["A", "A", "B", "B"]),
        ("Linus Wong", "S002", "92345678", ["", "A", "B", ""]),
    ]
    for row_index, (name, staff_id, mobile, codes) in enumerate(rows, start=7):
        ws.cell(row_index, 1, value=name)
        ws.cell(row_index, 3, value=staff_id)
        ws.cell(row_index, 4, value=mobile)
        for offset, code in enumerate(codes):
            ws.cell(row_index, 5 + offset, value=code)
    ws.cell(9, 1, value="Total")

    legend = [
        ("A", "9:45am-6:45pm (9 hr)"),
        ("B", "10:45am-8:15pm (9.5 hr)"),
    ]
    for offset, (code, text) in enumerate(legend, start=11):
        ws.cell(offset, 1, value=code)
        ws.cell(offset, 2, value=text)
    wb.save(path)
    return path


def save_d_and_g_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Applications"
    ws.cell(1, 2, value="Location: Harbour City")
    headers = [
        "Name",
        "Tel",
        "Staff ID",
        "Duties",
        "1 Apr",
        "2-Apr",
        "3 Apr",
        "1 May",
        "2 May",
        "Count",
        "Ttl Hours",
    ]
    for col, value in enumerate(headers, start=2):
        ws.cell(2, col, value=value)

    ws.cell(3, 2, value="Supervisor $95")
    staff_rows = [
        ("Fung King Yan Kelvin", "95153233", "S01245", "Supervisor", ["10:00-21:30", "", "", "10:30-23:00", ""]),
        ("Tam Kit Man Fion", "52401995", "S01280", "Supervisor", ["", "9:00-21:30", "10:00-21:30", "", ""]),
        ("LEE Sarah", "64364446", "S01060", "貨場", ["", "", "", "10:00-21:15", "9:00-21:30"]),
    ]
    for row, (name, tel, staff_id, duties, shifts) in enumerate(staff_rows, start=4):
        ws.cell(row, 2, value=name)
        ws.cell(row, 3, value=tel)
        ws.cell(row, 4, value=staff_id)
        ws.cell(row, 5, value=duties)
        for offset, shift in enumerate(shifts, start=6):
            ws.cell(row, offset, value=shift)

    ws.cell(7, 2, value="Helpers")
    ws.cell(8, 2, value="07:30-20:00")
    ws.cell(9, 2, value="Missing")
    ws.cell(11, 2, value="Slot Pattern")
    ws.cell(11, 3, value="Total Hours")
    ws.cell(12, 2, value="10:30-23:00")
    ws.cell(12, 3, value=12.5)
    wb.save(path)
    return path


def save_long_multi_month_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Apr - May"
    header_row = 6
    headers = ["Name", "Start Date", "Staff ID", "Mobile"]
    for index, value in enumerate(headers, start=1):
        ws.cell(header_row, index, value=value)
    start = date(2026, 4, 27)
    for offset in range(35):
        ws.cell(header_row, 5 + offset, value=start + timedelta(days=offset))
    ws.cell(header_row, 40, value="Days (May)")
    ws.cell(header_row, 41, value="Total Hours (May)")

    ws.cell(7, 1, value="Helper A")
    ws.cell(7, 3, value="S001")
    ws.cell(7, 4, value="91234567")
    for offset in range(35):
        ws.cell(7, 5 + offset, value="A" if offset in {0, 32, 33, 34} else "")
    ws.cell(7, 40, value=4)
    ws.cell(7, 41, value=36)
    ws.cell(8, 1, value="Total")

    ws.cell(10, 1, value="A")
    ws.cell(10, 2, value="9:45am-6:45pm (9 hr)")
    wb.save(path)
    return path


def save_slash_code_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Slash"
    header_row = 6
    for index, value in enumerate(["Name", "Start Date", "Staff ID", "Mobile"], start=1):
        ws.cell(header_row, index, value=value)
    for offset in range(4):
        ws.cell(header_row, 5 + offset, value=date(2025, 4, offset + 1))

    ws.cell(7, 1, value="John Chan")
    ws.cell(7, 3, value="S001")
    ws.cell(7, 4, value="91234567")
    for offset, code in enumerate(["A/C", "B/D", "A1/C1", "B1/D1"]):
        ws.cell(7, 5 + offset, value=code)
    ws.cell(8, 1, value="Total")

    ws.cell(10, 1, value="Tue - Sun")
    legend = [
        ("A", "9:45am-6:45pm (9 hr)"),
        ("C", "1:45pm-6:45pm (5 hr)"),
        ("B", "9:45am-8:15pm (10.5 hr)"),
        ("D", "9:45am-8:15pm (10.5 hr)"),
        ("A1", "1:45pm-6:45pm (5 hr)"),
        ("C1", "1:45pm-6:45pm (5 hr)"),
        ("B1", "1:45pm-8:15pm (6.5 hr)"),
        ("D1", "1:45pm-8:15pm (6.5 hr)"),
    ]
    for offset, (code, text) in enumerate(legend, start=11):
        ws.cell(offset, 1, value=code)
        ws.cell(offset, 2, value=text)
    wb.save(path)
    return path


def save_formula_header_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Formula"
    header_row = 6
    ws.cell(header_row, 1, value="Name")
    ws.cell(header_row, 2, value="Start Date")
    ws.cell(header_row, 3, value="Staff ID")
    ws.cell(header_row, 4, value="Mobile")
    ws.cell(header_row, 5, value=date(2025, 5, 1))
    ws.cell(header_row, 6, value="=E6+1")
    ws.cell(header_row, 7, value="=F6+1")
    ws.cell(7, 1, value="John Chan")
    ws.cell(7, 3, value="S001")
    ws.cell(7, 4, value="91234567")
    ws.cell(7, 5, value="A")
    ws.cell(7, 6, value="A")
    ws.cell(7, 7, value="A")
    ws.cell(8, 1, value="Total")
    ws.cell(10, 1, value="A")
    ws.cell(10, 2, value="9:45am-6:45pm (9 hr)")
    wb.save(path)
    return path
