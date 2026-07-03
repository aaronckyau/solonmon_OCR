from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .diagnostics import build_parse_diagnostics
from .excel_inspector import normalize_label
from .schema import (
    ParsedDateColumn,
    ParsedSchedule,
    ParsedScheduleEntry,
    ParsedStaff,
    ParserError,
    ParserWarning,
    to_jsonable,
)
from .time_utils import coerce_excel_date, format_time_minutes, parse_time_minutes


D_AND_G_PROJECT_PROFILE = "d_and_g"
D_AND_G_LAYOUT_TYPE = "d_and_g_job_applications"

SUMMARY_HEADERS = {
    "count",
    "ttl hours",
    "total hours",
    "salary",
    "client",
    "column1",
    "total around hours",
    "checking",
}
STAFF_STOP_MARKERS = {"missing", "slot pattern"}
MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}
MONTH_DATE_RE = re.compile(
    r"^\s*(\d{1,2})(?:st|nd|rd|th)?\s*[-/ ]\s*([A-Za-z]{3,9})\.?\s*$",
    re.IGNORECASE,
)
NUMERIC_DATE_RE = re.compile(r"^\s*(\d{1,2})\s*[-/]\s*(\d{1,2})\s*$")
TIME_RANGE_RE = re.compile(
    r"(?<!\d)(\d{1,2}(?::|\.)?\d{0,2}\s*(?:am|pm)?)\s*(?:-|to|~)\s*"
    r"(\d{1,2}(?::|\.)?\d{0,2}\s*(?:am|pm)?)(?!\d)",
    re.IGNORECASE,
)
YEAR_RE = re.compile(r"\b(20\d{2}|19\d{2})\b")


def parse_d_and_g_schedule(
    path_or_bytes: str | Path | bytes | bytearray | BytesIO,
    filename: str | None = None,
) -> ParsedSchedule:
    raw = _read_bytes(path_or_bytes)
    source_filename = filename or _filename(path_or_bytes)
    wb_formula = load_workbook(BytesIO(raw), data_only=False)
    wb_values = load_workbook(BytesIO(raw), data_only=True)
    warnings: list[ParserWarning] = []
    errors: list[ParserError] = []
    try:
        sheet_name = _select_sheet_name(wb_values)
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]
        header_row = _find_header_row(ws_values)
        if header_row is None:
            errors.append(ParserError(code="NO_D_AND_G_HEADER", message="Could not find the D&G Job Applications header row.", cell=None))
            return ParsedSchedule(
                project_profile=D_AND_G_PROJECT_PROFILE,
                source_filename=source_filename,
                sheet_name=sheet_name,
                layout_type=D_AND_G_LAYOUT_TYPE,
                header_row=None,
                warnings=warnings,
                errors=errors,
                diagnostics={
                    "layout_confidence": 0,
                    "year_source": "unavailable",
                },
            )

        header_cols = _detect_header_columns(ws_values, header_row)
        year, year_source = _infer_schedule_year(source_filename, ws_values)
        date_columns = _extract_date_columns(ws_formula, ws_values, header_row, year)
        if not date_columns:
            errors.append(ParserError(code="NO_DATE_COLUMNS", message="No D&G date columns were detected.", cell=None))
        staff = _extract_staff(ws_values, header_row, header_cols, date_columns)
        entries = _extract_entries(ws_values, header_row, staff, date_columns, header_cols)
        staff_end_row = max((item.row for item in staff), default=header_row)
        diagnostics = build_parse_diagnostics(
            selected_sheet_score=1,
            layout_confidence=0.95 if date_columns and staff else 0.4,
            header_row=header_row,
            staff_row_range={"start": staff[0].row if staff else None, "end": staff_end_row if staff else None},
            date_count=len(date_columns),
            staff_count=len(staff),
            entry_count=len(entries),
            entries=entries,
            warnings=warnings,
            errors=errors,
            ai_used=False,
        )
        diagnostics.update({
            "year": year,
            "year_source": year_source,
            "parser": D_AND_G_LAYOUT_TYPE,
        })
        return ParsedSchedule(
            project_profile=D_AND_G_PROJECT_PROFILE,
            source_filename=source_filename,
            sheet_name=sheet_name,
            layout_type=D_AND_G_LAYOUT_TYPE,
            header_row=header_row,
            date_columns=date_columns,
            staff=staff,
            shift_times={},
            entries=entries,
            warnings=warnings,
            errors=errors,
            diagnostics=to_jsonable(diagnostics),  # type: ignore[arg-type]
        )
    finally:
        wb_formula.close()
        wb_values.close()


def _select_sheet_name(workbook) -> str:
    for name in workbook.sheetnames:
        if "job applications" in name.lower():
            return name
    return workbook.active.title


def _find_header_row(ws_values) -> int | None:
    for row in range(1, min(ws_values.max_row, 25) + 1):
        labels = {normalize_label(ws_values.cell(row, col).value) for col in range(1, min(ws_values.max_column, 12) + 1)}
        if {"name", "tel", "staff id", "duties"}.issubset(labels):
            return row
    return None


def _detect_header_columns(ws_values, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for col in range(1, ws_values.max_column + 1):
        label = normalize_label(ws_values.cell(header_row, col).value)
        if label == "name":
            columns["name"] = col
        elif label in {"tel", "phone", "mobile"}:
            columns["phone"] = col
        elif label == "staff id":
            columns["staff_id"] = col
        elif label == "duties":
            columns["duties"] = col
    columns.setdefault("name", 2)
    columns.setdefault("phone", 3)
    columns.setdefault("staff_id", 4)
    columns.setdefault("duties", 5)
    return columns


def _extract_date_columns(ws_formula, ws_values, header_row: int, base_year: int) -> list[ParsedDateColumn]:
    columns: list[ParsedDateColumn] = []
    current_year = base_year
    last_month: int | None = None
    started = False
    for col in range(1, ws_values.max_column + 1):
        formula_value = ws_formula.cell(header_row, col).value
        cached_value = ws_values.cell(header_row, col).value
        label = normalize_label(cached_value if cached_value is not None else formula_value)
        if started and _is_summary_header(label):
            break

        parsed = coerce_excel_date(cached_value) or coerce_excel_date(formula_value)
        source = "excel_date" if parsed is not None else "text_month_header"
        if parsed is None:
            parsed = _parse_d_and_g_date_header(cached_value if cached_value is not None else formula_value, current_year)
        if parsed is None:
            if started and label:
                break
            continue

        if last_month is not None and parsed.month < last_month and parsed.year == current_year:
            current_year += 1
            parsed = parsed.replace(year=current_year)
        last_month = parsed.month
        started = True
        formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None
        columns.append(
            ParsedDateColumn(
                column=col,
                letter=get_column_letter(col),
                date=parsed.isoformat(),
                raw_value=to_jsonable(cached_value if cached_value is not None else formula_value),
                formula=formula,
                source=source,
            )
        )
    return columns


def _parse_d_and_g_date_header(value: Any, year: int) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    month_match = MONTH_DATE_RE.fullmatch(text)
    if month_match:
        day = int(month_match.group(1))
        month = MONTHS.get(month_match.group(2).lower())
        if month and 1 <= day <= 31:
            return date(year, month, day)
    numeric_match = NUMERIC_DATE_RE.fullmatch(text)
    if numeric_match:
        day = int(numeric_match.group(1))
        month = int(numeric_match.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return date(year, month, day)
    return None


def _is_summary_header(label: str) -> bool:
    if label in SUMMARY_HEADERS:
        return True
    return any(label.startswith(f"{header} ") for header in SUMMARY_HEADERS)


def _extract_staff(
    ws_values,
    header_row: int,
    header_cols: dict[str, int],
    date_columns: list[ParsedDateColumn],
) -> list[ParsedStaff]:
    staff: list[ParsedStaff] = []
    name_col = header_cols["name"]
    staff_id_col = header_cols["staff_id"]
    phone_col = header_cols["phone"]
    duties_col = header_cols["duties"]
    for row in range(header_row + 1, ws_values.max_row + 1):
        name = _cell_text(ws_values.cell(row, name_col).value)
        marker = normalize_label(name)
        if not name:
            continue
        if marker in STAFF_STOP_MARKERS or _looks_like_time_range(name):
            break

        staff_id = _cell_text(ws_values.cell(row, staff_id_col).value)
        phone_last4 = _phone_last4(ws_values.cell(row, phone_col).value)
        has_schedule = any(_parse_d_and_g_time_range(ws_values.cell(row, date_col.column).value) for date_col in date_columns)
        if not staff_id and not phone_last4 and not has_schedule:
            continue

        staff.append(
            ParsedStaff(
                name=name,
                staff_id=staff_id,
                phone_last4=phone_last4,
                row=row,
                source_cells={
                    "name": ws_values.cell(row, name_col).coordinate,
                    "staff_id": ws_values.cell(row, staff_id_col).coordinate,
                    "phone_last4": ws_values.cell(row, phone_col).coordinate,
                    "duties": ws_values.cell(row, duties_col).coordinate,
                },
            )
        )
    return staff


def _extract_entries(
    ws_values,
    header_row: int,
    staff: list[ParsedStaff],
    date_columns: list[ParsedDateColumn],
    header_cols: dict[str, int],
) -> list[ParsedScheduleEntry]:
    entries: list[ParsedScheduleEntry] = []
    for staff_item in staff:
        for date_col in date_columns:
            cell = ws_values.cell(staff_item.row, date_col.column)
            raw_value = cell.value
            if raw_value in (None, ""):
                continue
            raw_text = _cell_text(raw_value)
            if not raw_text:
                continue
            parsed = _parse_d_and_g_time_range(raw_text)
            warnings: list[str] = []
            scheduled_in = ""
            scheduled_out = ""
            scheduled_hours: float | None = None
            resolution_source = "d_and_g_time_range"
            if parsed:
                scheduled_in = parsed["start"]
                scheduled_out = parsed["end"]
                scheduled_hours = parsed["hours"]
            else:
                warnings.append("UNPARSED_D_AND_G_TIME_RANGE")
                resolution_source = "unresolved"

            entries.append(
                ParsedScheduleEntry(
                    staff_name=staff_item.name,
                    staff_id=staff_item.staff_id,
                    phone_last4=staff_item.phone_last4,
                    staff_row=staff_item.row,
                    date=date_col.date,
                    date_column=date_col.column,
                    date_cell=f"{date_col.letter}{header_row}",
                    schedule_cell=cell.coordinate,
                    raw_value=to_jsonable(raw_value),
                    raw_shift_code=raw_text,
                    shift_code="",
                    scheduled_in=scheduled_in,
                    scheduled_out=scheduled_out,
                    scheduled_hours=scheduled_hours,
                    resolution_source=resolution_source,
                    warnings=warnings,
                    shift_options=[],
                )
            )
    return entries


def _parse_d_and_g_time_range(value: Any) -> dict[str, Any] | None:
    source = str(value or "").strip()
    if not source:
        return None
    match = TIME_RANGE_RE.search(source.replace("–", "-").replace("—", "-"))
    if not match:
        return None
    start_minutes = parse_time_minutes(match.group(1))
    end_minutes = parse_time_minutes(match.group(2))
    if start_minutes is None or end_minutes is None:
        return None
    delta = end_minutes - start_minutes
    if delta <= 0:
        delta += 24 * 60
    if delta > 24 * 60:
        return None
    return {
        "start": format_time_minutes(start_minutes),
        "end": format_time_minutes(end_minutes),
        "hours": round(delta / 60, 2),
        "source_text": source,
    }


def _infer_schedule_year(source_filename: str, ws_values) -> tuple[int, str]:
    filename_match = YEAR_RE.search(source_filename)
    if filename_match:
        return int(filename_match.group(1)), "filename"
    for row in range(1, min(ws_values.max_row, 20) + 1):
        for col in range(1, min(ws_values.max_column, 12) + 1):
            value = ws_values.cell(row, col).value
            if isinstance(value, datetime):
                return value.year, "workbook_date"
            if isinstance(value, date):
                return value.year, "workbook_date"
            match = YEAR_RE.search(str(value or ""))
            if match:
                return int(match.group(1)), "workbook_text"
    return datetime.today().year, "current_year"


def _looks_like_time_range(value: Any) -> bool:
    return TIME_RANGE_RE.search(str(value or "")) is not None


def _phone_last4(value: Any) -> str:
    digits = re.sub(r"\D+", "", str(value or ""))
    return digits[-4:] if len(digits) >= 4 else digits


def _cell_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _read_bytes(path_or_bytes: str | Path | bytes | bytearray | BytesIO) -> bytes:
    if isinstance(path_or_bytes, bytes):
        return path_or_bytes
    if isinstance(path_or_bytes, bytearray):
        return bytes(path_or_bytes)
    if isinstance(path_or_bytes, BytesIO):
        return path_or_bytes.getvalue()
    return Path(path_or_bytes).read_bytes()


def _filename(path_or_bytes: Any) -> str:
    if isinstance(path_or_bytes, (str, Path)):
        return Path(path_or_bytes).name
    return "schedule.xlsx"
