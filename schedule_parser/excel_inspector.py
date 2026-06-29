from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .schema import to_jsonable
from .time_utils import coerce_excel_date


HEADER_LABELS = {
    "name",
    "staff name",
    "employee name",
    "staff",
    "姓名",
    "員工姓名",
    "雇員姓名",
}

STAFF_RELATED_LABELS = {
    "staff id",
    "staffid",
    "id",
    "mobile",
    "phone",
    "tel",
    "telephone",
    "start date",
    "startdate",
}


def inspect_workbook(path_or_bytes: str | Path | bytes | bytearray | BytesIO, filename: str | None = None) -> dict[str, Any]:
    raw = _read_bytes(path_or_bytes)
    source_filename = filename or _filename(path_or_bytes)
    wb_formula = load_workbook(BytesIO(raw), data_only=False)
    wb_values = load_workbook(BytesIO(raw), data_only=True)
    try:
        sheets = []
        for sheet_name in wb_formula.sheetnames:
            sheets.append(_inspect_sheet(wb_formula[sheet_name], wb_values[sheet_name]))
        return {
            "filename": source_filename,
            "sheet_names": wb_formula.sheetnames,
            "active_sheet_name": wb_formula.active.title,
            "sheets": sheets,
        }
    finally:
        wb_formula.close()
        wb_values.close()


def _inspect_sheet(ws_formula, ws_values) -> dict[str, Any]:
    header_rows = _candidate_header_rows(ws_values)
    date_columns = []
    for header in header_rows:
        date_columns.extend(_candidate_date_columns(ws_formula, ws_values, header["row"]))
    return {
        "sheet_name": ws_formula.title,
        "max_row": ws_formula.max_row,
        "max_column": ws_formula.max_column,
        "merged_ranges": [str(rng) for rng in ws_formula.merged_cells.ranges],
        "first_20_rows": _first_rows(ws_values, 20, 20),
        "formula_cells": _formula_cells(ws_formula, 80, 80),
        "candidate_header_rows": header_rows,
        "candidate_date_columns": date_columns,
        "non_empty_cell_density": _non_empty_density(ws_values),
        "appears_old_or_archive": _appears_old_or_archive(ws_formula.title),
        "used_range": _used_range(ws_values),
    }


def _candidate_header_rows(ws) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for row in range(1, min(ws.max_row, 30) + 1):
        labels: dict[str, int] = {}
        score = 0
        for col in range(1, min(ws.max_column, 40) + 1):
            raw = ws.cell(row, col).value
            normalized = normalize_label(raw)
            if not normalized:
                continue
            if normalized in HEADER_LABELS or normalized.endswith(" name") or "name" == normalized:
                labels["name"] = col
                score += 5
            if normalized in STAFF_RELATED_LABELS:
                labels[normalized] = col
                score += 2
            elif normalized.replace(" ", "") in STAFF_RELATED_LABELS:
                labels[normalized.replace(" ", "")] = col
                score += 2
        if "name" in labels:
            candidates.append({"row": row, "score": score, "labels": labels})
    return candidates


def _candidate_date_columns(ws_formula, ws_values, header_row: int) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    last_date = None
    last_col = None
    for col in range(1, ws_formula.max_column + 1):
        formula_cell = ws_formula.cell(header_row, col)
        value_cell = ws_values.cell(header_row, col)
        formula_value = formula_cell.value
        cached_value = value_cell.value
        parsed_date = coerce_excel_date(cached_value) or coerce_excel_date(formula_value)
        source = "cached_value" if coerce_excel_date(cached_value) else "direct_value"
        formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None
        if parsed_date is None and formula and last_date is not None and _is_previous_plus_one_formula(formula, header_row, last_col):
            parsed_date = last_date + __import__("datetime").timedelta(days=1)
            source = "formula_inferred"
        if parsed_date is None:
            continue
        columns.append(
            {
                "row": header_row,
                "column": col,
                "letter": get_column_letter(col),
                "date": parsed_date.isoformat(),
                "raw_value": to_jsonable(cached_value if cached_value is not None else formula_value),
                "formula": formula,
                "source": source,
            }
        )
        last_date = parsed_date
        last_col = col
    return columns


def normalize_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _first_rows(ws, row_count: int, col_count: int) -> list[list[Any]]:
    rows = []
    for row in range(1, min(ws.max_row, row_count) + 1):
        rows.append([to_jsonable(ws.cell(row, col).value) for col in range(1, min(ws.max_column, col_count) + 1)])
    return rows


def _formula_cells(ws, row_count: int, col_count: int) -> list[dict[str, Any]]:
    formulas = []
    for row in range(1, min(ws.max_row, row_count) + 1):
        for col in range(1, min(ws.max_column, col_count) + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.startswith("="):
                formulas.append({"cell": ws.cell(row, col).coordinate, "formula": value})
    return formulas


def _non_empty_density(ws) -> float:
    max_row = max(ws.max_row, 1)
    max_column = max(ws.max_column, 1)
    total = max_row * max_column
    non_empty = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                non_empty += 1
    return round(non_empty / total, 4)


def _used_range(ws) -> dict[str, int | None]:
    min_row = min_col = None
    max_row = max_col = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_row = cell.row if max_row is None else max(max_row, cell.row)
            max_col = cell.column if max_col is None else max(max_col, cell.column)
    return {"min_row": min_row, "min_column": min_col, "max_row": max_row, "max_column": max_col}


def _appears_old_or_archive(sheet_name: str) -> bool:
    normalized = sheet_name.strip().lower()
    return normalized.endswith("old") or "archive" in normalized or "old" in normalized


def _is_previous_plus_one_formula(formula: str, row: int, previous_col: int | None) -> bool:
    if previous_col is None:
        return False
    match = re.fullmatch(r"=\s*\$?([A-Z]{1,3})\$?(\d+)\s*\+\s*1\s*", formula.strip(), re.IGNORECASE)
    if not match:
        return False
    ref_col = match.group(1).upper()
    ref_row = int(match.group(2))
    return ref_row == row and ref_col == get_column_letter(previous_col)


def _read_bytes(path_or_bytes: str | Path | bytes | bytearray | BytesIO) -> bytes:
    if isinstance(path_or_bytes, bytes):
        return path_or_bytes
    if isinstance(path_or_bytes, bytearray):
        return bytes(path_or_bytes)
    if isinstance(path_or_bytes, BytesIO):
        return path_or_bytes.getvalue()
    return Path(path_or_bytes).read_bytes()


def _filename(path_or_bytes: Any) -> str:
    if isinstance(path_or_bytes, (str, Path)):
        return Path(path_or_bytes).name
    return "workbook.xlsx"
