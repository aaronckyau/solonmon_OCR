from __future__ import annotations

import argparse
import json
import re
from datetime import timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .diagnostics import build_parse_diagnostics
from .excel_inspector import inspect_workbook, normalize_label
from .layout_detector import (
    DIRECT_TIME_WITH_HOURS_COLUMNS,
    FIRST_SHEET_SELECTED_OVER_HIGHER_SCORING_SHEET,
    SHIFT_CODE_MATRIX_WITH_LEGEND,
    detect_schedule_layout,
)
from .rule_resolver import (
    ShiftRuleAIResolver,
    is_section_header,
    normalize_shift_code,
    parse_direct_time_cell,
    parse_legend_from_sheet,
    resolve_shift_code,
)
from .schema import (
    ParsedDateColumn,
    ParsedSchedule,
    ParsedScheduleEntry,
    ParsedShiftTime,
    ParsedStaff,
    ParserError,
    ParserWarning,
    to_jsonable,
)
from .time_utils import coerce_excel_date


SUMMARY_HEADERS = {
    "days",
    "total hours",
    "normal days total hours",
    "normal total hours",
    "deduct hours",
    "normal hours",
    "ph total hours",
    "holiday total hours",
    "salary",
    "client",
    "salary total",
    "client total",
}

STAFF_STOP_MARKERS = {"need", "total", "briefing", "still need"}
MAX_DATE_COLUMNS = 370


def parse_oil_street_schedule(
    path_or_bytes: str | Path | bytes | bytearray | BytesIO,
    filename: str | None = None,
    ai_resolver: ShiftRuleAIResolver | None = None,
) -> ParsedSchedule:
    raw = _read_bytes(path_or_bytes)
    source_filename = filename or _filename(path_or_bytes)
    inspected = inspect_workbook(raw, filename=source_filename)
    detection = detect_schedule_layout(inspected)
    warnings = [_layout_warning(code, detection.diagnostics) for code in detection.warnings]
    errors: list[ParserError] = []
    entries: list[ParsedScheduleEntry] = []
    staff: list[ParsedStaff] = []
    date_columns: list[ParsedDateColumn] = []
    shift_times: dict[str, ParsedShiftTime] = {}
    ai_used = False

    if not detection.sheet_name or not detection.header_row:
        errors.append(ParserError(code="NO_SCHEDULE_SHEET", message="Could not detect a schedule sheet/header row.", cell=None))
        return ParsedSchedule(
            project_profile="oil_street",
            source_filename=source_filename,
            sheet_name=detection.sheet_name or "",
            layout_type=detection.layout_type,
            header_row=detection.header_row,
            warnings=warnings,
            errors=errors,
            diagnostics={"inspector": inspected, "layout_detection": detection.to_dict()},
        )

    wb_formula = load_workbook(BytesIO(raw), data_only=False)
    wb_values = load_workbook(BytesIO(raw), data_only=True)
    try:
        ws_formula = wb_formula[detection.sheet_name]
        ws_values = wb_values[detection.sheet_name]
        header_row = detection.header_row
        layout_type = detection.layout_type

        header_cols = _detect_header_columns(ws_values, header_row)
        date_columns = _extract_date_columns(ws_formula, ws_values, header_row, layout_type)
        if not date_columns:
            errors.append(ParserError(code="NO_DATE_COLUMNS", message="No date columns were detected.", cell=None))

        staff = _extract_staff(ws_values, header_row, header_cols)
        staff_end_row = max((item.row for item in staff), default=header_row)
        if layout_type == SHIFT_CODE_MATRIX_WITH_LEGEND:
            shift_times = parse_legend_from_sheet(ws_formula, ws_values, header_row, staff_end_row)
            if not shift_times and ai_resolver is not None:
                ai_payload = ai_resolver.resolve(_compact_ai_context(ws_values, header_row, staff_end_row))
                shift_times = _shift_times_from_ai(ai_payload)
                ai_used = bool(shift_times)

        entries = _extract_entries(
            ws_values=ws_values,
            staff=staff,
            date_columns=date_columns,
            header_row=header_row,
            layout_type=layout_type,
            shift_times=shift_times,
        )
        for entry in entries:
            for code in entry.warnings:
                warnings.append(
                    ParserWarning(
                        code=code,
                        message=_warning_message(code, entry.shift_code),
                        cell=entry.schedule_cell,
                        severity="warning",
                    )
                )

        diagnostics = build_parse_diagnostics(
            selected_sheet_score=detection.diagnostics.get("overall_score"),
            layout_confidence=detection.confidence,
            header_row=header_row,
            staff_row_range={"start": staff[0].row if staff else None, "end": staff_end_row if staff else None},
            date_count=len(date_columns),
            staff_count=len(staff),
            entry_count=len(entries),
            entries=entries,
            warnings=warnings,
            errors=errors,
            ai_used=ai_used,
        )
        diagnostics["layout_detection"] = detection.to_dict()
        diagnostics["workbook"] = {
            "sheet_names": inspected.get("sheet_names", []),
            "active_sheet_name": inspected.get("active_sheet_name"),
        }

        return ParsedSchedule(
            project_profile="oil_street",
            source_filename=source_filename,
            sheet_name=detection.sheet_name,
            layout_type=layout_type,
            header_row=header_row,
            date_columns=date_columns,
            staff=staff,
            shift_times=shift_times,
            entries=entries,
            warnings=warnings,
            errors=errors,
            diagnostics=to_jsonable(diagnostics),  # type: ignore[arg-type]
        )
    finally:
        wb_formula.close()
        wb_values.close()


def _detect_header_columns(ws_values, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for col in range(1, ws_values.max_column + 1):
        label = normalize_label(ws_values.cell(header_row, col).value)
        compact = label.replace(" ", "")
        if label in {"name", "staff name", "employee name"} or label.endswith(" name"):
            columns["name"] = col
        elif label in {"staff id", "id"} or compact in {"staffid", "employeeid"}:
            columns["staff_id"] = col
        elif label in {"mobile", "phone", "tel", "telephone"}:
            columns["mobile"] = col
        elif label == "start date" or compact == "startdate":
            columns["start_date"] = col
    columns.setdefault("name", 1)
    columns.setdefault("staff_id", 3)
    columns.setdefault("mobile", 4)
    return columns


def _extract_date_columns(ws_formula, ws_values, header_row: int, layout_type: str) -> list[ParsedDateColumn]:
    columns: list[ParsedDateColumn] = []
    last_date = None
    last_col = None
    non_date_after_dates = 0
    for col in range(1, ws_formula.max_column + 1):
        formula_value = ws_formula.cell(header_row, col).value
        cached_value = ws_values.cell(header_row, col).value
        header_label = normalize_label(cached_value if cached_value is not None else formula_value)
        if columns and _is_summary_header(header_label):
            break

        cached_date = coerce_excel_date(cached_value)
        direct_date = coerce_excel_date(formula_value)
        parsed_date = cached_date or direct_date
        source = "cached_value" if cached_date else "direct_value"
        formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None
        if parsed_date is None and formula and last_date is not None and _previous_plus_one(formula, header_row, last_col):
            parsed_date = last_date + timedelta(days=1)
            source = "formula_inferred"

        if parsed_date is None:
            if columns:
                non_date_after_dates += 1
                if layout_type == SHIFT_CODE_MATRIX_WITH_LEGEND:
                    if header_label or non_date_after_dates >= 2:
                        break
                elif layout_type == DIRECT_TIME_WITH_HOURS_COLUMNS and len(columns) >= 28 and non_date_after_dates >= 2:
                    break
            continue

        non_date_after_dates = 0
        columns.append(
            ParsedDateColumn(
                column=col,
                letter=get_column_letter(col),
                date=parsed_date.isoformat(),
                raw_value=to_jsonable(cached_value if cached_value is not None else formula_value),
                formula=formula,
                source=source,
            )
        )
        last_date = parsed_date
        last_col = col
        if len(columns) >= MAX_DATE_COLUMNS:
            break
    return columns


def _is_summary_header(label: str) -> bool:
    if label in SUMMARY_HEADERS:
        return True
    return any(label.startswith(f"{header} ") or label.startswith(f"{header} (") for header in SUMMARY_HEADERS)


def _extract_staff(ws_values, header_row: int, header_cols: dict[str, int]) -> list[ParsedStaff]:
    staff: list[ParsedStaff] = []
    blank_streak = 0
    started = False
    name_col = header_cols["name"]
    staff_id_col = header_cols["staff_id"]
    mobile_col = header_cols["mobile"]
    for row in range(header_row + 1, ws_values.max_row + 1):
        raw_name = ws_values.cell(row, name_col).value
        name = str(raw_name or "").strip()
        marker = name.lower()
        if marker in STAFF_STOP_MARKERS or marker.startswith("briefing"):
            break
        if started and is_section_header(name):
            break
        if not name:
            if started:
                blank_streak += 1
                if blank_streak >= 3:
                    break
            continue
        started = True
        blank_streak = 0
        staff_id = str(ws_values.cell(row, staff_id_col).value or "").strip()
        phone_last4 = _phone_last4(ws_values.cell(row, mobile_col).value)
        staff.append(
            ParsedStaff(
                name=name,
                staff_id=staff_id,
                phone_last4=phone_last4,
                row=row,
                source_cells={
                    "name": ws_values.cell(row, name_col).coordinate,
                    "staff_id": ws_values.cell(row, staff_id_col).coordinate,
                    "phone_last4": ws_values.cell(row, mobile_col).coordinate,
                },
            )
        )
    return staff


def _extract_entries(
    *,
    ws_values,
    staff: list[ParsedStaff],
    date_columns: list[ParsedDateColumn],
    header_row: int,
    layout_type: str,
    shift_times: dict[str, ParsedShiftTime],
) -> list[ParsedScheduleEntry]:
    entries: list[ParsedScheduleEntry] = []
    for staff_item in staff:
        for date_col in date_columns:
            cell = ws_values.cell(staff_item.row, date_col.column)
            raw_value = cell.value
            if raw_value in (None, ""):
                continue
            raw_text = str(raw_value).strip()
            if not raw_text:
                continue
            if layout_type == DIRECT_TIME_WITH_HOURS_COLUMNS:
                adjacent_hours = ws_values.cell(staff_item.row, date_col.column + 1).value
                resolution = parse_direct_time_cell(raw_text, adjacent_hours)
                if resolution is None:
                    code = normalize_shift_code(raw_text)
                    resolution = {
                        "shift_code": code,
                        "scheduled_in": "",
                        "scheduled_out": "",
                        "scheduled_hours": None,
                        "resolution_source": "unresolved",
                        "warnings": ["UNPARSED_DIRECT_TIME"],
                    }
            else:
                resolution = resolve_shift_code(raw_text, date_col.date, shift_times)
                if resolution.get("resolution_source") == "unresolved":
                    direct = parse_direct_time_cell(raw_text)
                    resolution = direct or resolution

            entries.append(
                ParsedScheduleEntry(
                    staff_name=staff_item.name,
                    staff_id=staff_item.staff_id,
                    phone_last4=staff_item.phone_last4,
                    staff_row=staff_item.row,
                    date=date_col.date,
                    date_column=date_col.column,
                    date_cell=f"{date_col.letter}{header_row}",
                    schedule_cell=cell.coordinate,
                    raw_value=to_jsonable(raw_value),
                    raw_shift_code=raw_text,
                    shift_code=resolution.get("shift_code") or normalize_shift_code(raw_text),
                    scheduled_in=resolution.get("scheduled_in") or "",
                    scheduled_out=resolution.get("scheduled_out") or "",
                    scheduled_hours=resolution.get("scheduled_hours"),
                    resolution_source=resolution.get("resolution_source") or "unresolved",
                    warnings=list(resolution.get("warnings") or []),
                    shift_options=list(resolution.get("shift_options") or []),
                )
            )
    return entries


def _compact_ai_context(ws_values, header_row: int, staff_end_row: int) -> dict[str, Any]:
    rows = []
    for row in range(max(header_row + 1, staff_end_row + 1), ws_values.max_row + 1):
        values = [to_jsonable(ws_values.cell(row, col).value) for col in range(1, ws_values.max_column + 1)]
        if any(value not in (None, "") for value in values):
            rows.append(values)
    return {"header_row": header_row, "legend_rows": rows}


def _shift_times_from_ai(payload: dict[str, Any]) -> dict[str, ParsedShiftTime]:
    shift_times: dict[str, ParsedShiftTime] = {}
    for item in payload.get("shift_times", []) if isinstance(payload, dict) else []:
        code = normalize_shift_code(item.get("code"))
        start = item.get("start") or ""
        end = item.get("end") or ""
        if not code or not start or not end:
            continue
        shift_times[code] = ParsedShiftTime(
            code=code,
            start=start,
            end=end,
            hours=item.get("hours"),
            source="ai",
            source_text=item.get("source_text") or "AI parsed",
            source_cell=None,
            applies_to=item.get("applies_to") or "default",
            specific_dates=list(item.get("specific_dates") or []),
        )
    return shift_times


def _warning_message(code: str, shift_code: str) -> str:
    if code == "UNKNOWN_SHIFT_CODE":
        return f"Unknown shift code: {shift_code}"
    if code == "AMBIGUOUS_SHIFT_CODE":
        return f"Ambiguous shift code: {shift_code}"
    if code == "UNKNOWN_OR_SHIFT_PART":
        return f"OR shift code contains unknown part: {shift_code}"
    if code == "UNPARSED_DIRECT_TIME":
        return "Could not parse direct time range."
    return code


def _layout_warning(code: str, diagnostics: dict[str, Any]) -> ParserWarning:
    return ParserWarning(
        code=code,
        message=_layout_warning_message(code, diagnostics),
        cell=None,
        severity="warning",
    )


def _layout_warning_message(code: str, diagnostics: dict[str, Any]) -> str:
    if code == FIRST_SHEET_SELECTED_OVER_HIGHER_SCORING_SHEET:
        selected = diagnostics.get("selected_first_sheet") or diagnostics.get("sheet_name")
        ignored = diagnostics.get("ignored_higher_scoring_sheet") or {}
        ignored_name = ignored.get("sheet_name") if isinstance(ignored, dict) else None
        if selected and ignored_name:
            return (
                f"已優先讀取第一個工作表「{selected}」；另有工作表「{ignored_name}」偵測分數較高但已忽略，"
                "請確認目前使用的工作表正確。"
            )
        return "已優先讀取第一個可解析工作表；請確認目前使用的工作表正確。"
    return code


def _previous_plus_one(formula: str, row: int, previous_col: int | None) -> bool:
    if previous_col is None:
        return False
    match = re.fullmatch(r"=\s*\$?([A-Z]{1,3})\$?(\d+)\s*\+\s*1\s*", formula.strip(), re.IGNORECASE)
    if not match:
        return False
    return int(match.group(2)) == row and match.group(1).upper() == get_column_letter(previous_col)


def _phone_last4(value: Any) -> str:
    digits = re.sub(r"\D+", "", str(value or ""))
    return digits[-4:] if len(digits) >= 4 else digits


def _read_bytes(path_or_bytes: str | Path | bytes | bytearray | BytesIO) -> bytes:
    if isinstance(path_or_bytes, bytes):
        return path_or_bytes
    if isinstance(path_or_bytes, bytearray):
        return bytes(path_or_bytes)
    if isinstance(path_or_bytes, BytesIO):
        return path_or_bytes.getvalue()
    return Path(path_or_bytes).read_bytes()


def _filename(path_or_bytes: Any) -> str:
    if isinstance(path_or_bytes, (str, Path)):
        return Path(path_or_bytes).name
    return "schedule.xlsx"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Parse Oil Street schedule Excel files.")
    parser.add_argument("path", help="Path to .xlsx or .xlsm schedule file")
    parser.add_argument("--json", action="store_true", help="Print full ParsedSchedule as formatted JSON")
    parser.add_argument("--summary", action="store_true", help="Print parser summary")
    args = parser.parse_args(argv)

    parsed = parse_oil_street_schedule(args.path)
    data = parsed.to_dict()
    if args.json or not args.summary:
        print(json.dumps(data, ensure_ascii=False, indent=2))
        return 0

    diagnostics = data["diagnostics"]
    print(f"selected sheet: {data['sheet_name']}")
    print(f"layout type: {data['layout_type']}")
    print(f"header row: {data['header_row']}")
    print(f"date count: {diagnostics['date_count']}")
    print(f"staff count: {diagnostics['staff_count']}")
    print(f"entry count: {diagnostics['entry_count']}")
    print(f"shift code count: {len(data['shift_times'])}")
    print(f"unknown shift codes: {', '.join(diagnostics['unknown_shift_codes']) or '-'}")
    print(f"warnings: {diagnostics['warning_count']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
